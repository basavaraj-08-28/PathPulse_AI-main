"""
PathPulse AI - Pathole Detection & Mapping System
Backend API built with Flask
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime, timezone
import os
import io
import csv
import shutil

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(basedir, 'templates'),
    static_folder=os.path.join(basedir, 'static')
)
CORS(app)

# Database configuration
# 1. Turso Database (libSQL) via TURSO_DATABASE_URL and TURSO_AUTH_TOKEN
DEFAULT_TURSO_URL = 'libsql://pathpulse-db-basavaraj-08-28.aws-ap-south-1.turso.io'
DEFAULT_TURSO_TOKEN = 'eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJhIjoicnciLCJpYXQiOjE3ODgxOTg2OTMsImlkIjoiMDFhMDU4ZjItMmMwMS03MDkzLTkyNmItMTY5YTgxN2VmZTk0Iiwia2lkIjoiUGluVE5KblVTMWJLcnhUZGVsTVdqeHZFUUo0cTZDN2pOR2twdlN6MG9DMCIsInJpZCI6IjAzMjZjMjRhLTJhYjctNDE4Zi04YzVmLTFlYjUwNDZkYmI3YSJ9.XPeaaVZQw27FWF_wt9wsjCYIe93B5iV9J0yT1pj3D3z8ATKPJdjitPoQv5g-odwev-Q4WPzaNQXSevJu7liZDg'

turso_url = os.getenv('TURSO_DATABASE_URL') or os.getenv('TURSO_URL') or DEFAULT_TURSO_URL
turso_token = os.getenv('TURSO_AUTH_TOKEN') or DEFAULT_TURSO_TOKEN
db_url = os.getenv('DATABASE_URL') or os.getenv('POSTGRES_URL')

if db_url and not db_url.startswith("libsql://"):
    if db_url.startswith("postgres://"):
        # Modern SQLAlchemy requires postgresql:// instead of legacy postgres://
        app.config['SQLALCHEMY_DATABASE_URI'] = db_url.replace("postgres://", "postgresql://", 1)
    else:
        app.config['SQLALCHEMY_DATABASE_URI'] = db_url
else:
    # Helper to check if a directory is writable
    def _is_dir_writable(path):
        try:
            testfile = os.path.join(path, '.perm_test')
            with open(testfile, 'w') as f:
                f.write('1')
            os.remove(testfile)
            return True
        except Exception:
            return False

    # Check if running in serverless environment (Vercel, AWS Lambda, etc.) or read-only filesystem
    is_serverless = bool(
        os.getenv('VERCEL') or
        os.getenv('VERCEL_ENV') or
        os.getenv('VERCEL_URL') or
        os.getenv('VERCEL_REGION') or
        os.getenv('NOW_REGION') or
        os.getenv('AWS_LAMBDA_FUNCTION_NAME') or
        os.getenv('LAMBDA_TASK_ROOT') or
        not _is_dir_writable(basedir)
    )

    if is_serverless:
        tmp_db = '/tmp/pathpulse.db'
        src_db = os.path.join(basedir, 'pathpulse.db')
        # Copy initial database file to /tmp if it does not exist yet so seed data is preserved
        if not os.path.exists(tmp_db) and os.path.exists(src_db):
            try:
                # Use copyfile to avoid copying read-only file metadata from Vercel package
                shutil.copyfile(src_db, tmp_db)
                os.chmod(tmp_db, 0o666)
            except Exception as e:
                print(f"[PathPulse] Warning: could not copy initial DB to /tmp: {e}")
        elif os.path.exists(tmp_db):
            try:
                os.chmod(tmp_db, 0o666)
            except Exception:
                pass
        app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{tmp_db}'
    else:
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'pathpulse.db')

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'pathpulse-secret-key-2026')

db = SQLAlchemy(app)

def sync_to_turso(sql, params=None):
    """Sync database mutations directly to Turso Edge cloud database"""
    if not (turso_url and turso_token):
        return
    try:
        clean_url = turso_url.replace("libsql://", "https://")
        if not clean_url.startswith("https://"):
            clean_url = f"https://{clean_url}"
        pipeline_url = f"{clean_url.rstrip('/')}/v2/pipeline"

        stmt = {"sql": sql}
        if params:
            args = []
            for p in params:
                if p is None:
                    args.append({"type": "null"})
                elif isinstance(p, (int, bool)):
                    args.append({"type": "integer", "value": str(int(p))})
                elif isinstance(p, float):
                    args.append({"type": "float", "value": p})
                else:
                    args.append({"type": "text", "value": str(p)})
            stmt["args"] = args

        payload = {"requests": [{"type": "execute", "stmt": stmt}, {"type": "close"}]}
        headers = {"Authorization": f"Bearer {turso_token}", "Content-Type": "application/json"}
        import requests
        requests.post(pipeline_url, headers=headers, json=payload, timeout=4)
    except Exception as e:
        print(f"[PathPulse] Turso sync notice: {e}")



# ─── Models ───────────────────────────────────────────────────────────────────

class Pathole(db.Model):
    """Stores detected pathole locations"""
    id = db.Column(db.Integer, primary_key=True)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    severity = db.Column(db.String(20), nullable=False, default='medium')  # low, medium, high
    confidence = db.Column(db.Float, nullable=False, default=0.5)
    reported_by = db.Column(db.String(100), default='anonymous')
    report_count = db.Column(db.Integer, default=1)
    accel_peak = db.Column(db.Float, nullable=True)  # peak acceleration value
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc),
                           onupdate=lambda: datetime.now(timezone.utc))
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return {
            'id': self.id,
            'latitude': self.latitude,
            'longitude': self.longitude,
            'severity': self.severity,
            'confidence': self.confidence,
            'reported_by': self.reported_by,
            'report_count': self.report_count,
            'accel_peak': self.accel_peak,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None,
            'is_active': self.is_active
        }


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Serve the main application page"""
    return render_template('index.html')


@app.route('/map')
def map_page():
    """Serve the dedicated map page"""
    return render_template('map.html')


@app.route('/detect')
def detect():
    """Serve the detection/ride mode page (Admin only)"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('map_page'))
    return render_template('detect.html')


@app.route('/admin')
def admin_page():
    """Serve the admin portal page"""
    return render_template('admin.html')


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Authenticate admin credentials and set session"""
    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')

    expected_user = os.getenv('ADMIN_USER', 'basavarajyn123@gmail.com')
    expected_pass = os.getenv('ADMIN_PASSWORD', 'admin123')

    if username == expected_user and password == expected_pass:
        session['admin_logged_in'] = True
        return jsonify({'status': 'success', 'message': 'Logged in successfully'})
    return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401


@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """Clear admin session"""
    session.pop('admin_logged_in', None)
    return jsonify({'status': 'success', 'message': 'Logged out successfully'})


@app.route('/api/admin/check', methods=['GET'])
def admin_check():
    """Verify if admin session is active"""
    is_logged_in = session.get('admin_logged_in', False)
    return jsonify({'logged_in': is_logged_in})


@app.route('/api/patholes', methods=['GET'])
def get_patholes():
    """Get patholes, optionally filtered by bounds and active status"""
    lat_min = request.args.get('lat_min', type=float)
    lat_max = request.args.get('lat_max', type=float)
    lng_min = request.args.get('lng_min', type=float)
    lng_max = request.args.get('lng_max', type=float)
    status = request.args.get('status', 'active')  # active, resolved, all

    if status == 'active':
        query = Pathole.query.filter_by(is_active=True)
    elif status == 'resolved':
        query = Pathole.query.filter_by(is_active=False)
    else:
        query = Pathole.query

    if all(v is not None for v in [lat_min, lat_max, lng_min, lng_max]):
        query = query.filter(
            Pathole.latitude.between(lat_min, lat_max),
            Pathole.longitude.between(lng_min, lng_max)
        )

    patholes = query.order_by(Pathole.created_at.desc()).all()
    return jsonify({
        'status': 'success',
        'count': len(patholes),
        'patholes': [p.to_dict() for p in patholes]
    })


@app.route('/api/patholes', methods=['POST'])
def report_pathole():
    """Report a new pathole detected by accelerometer or manual report (Admin only)"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Forbidden: Only admin is authorized to detect and record potholes.'}), 403

    data = request.get_json()

    if not data or 'latitude' not in data or 'longitude' not in data:
        return jsonify({'status': 'error', 'message': 'latitude and longitude are required'}), 400

    try:
        lat = float(data['latitude'])
        lng = float(data['longitude'])
    except (ValueError, TypeError):
        return jsonify({'status': 'error', 'message': 'Invalid latitude or longitude coordinates'}), 400

    try:
        # Check if a pathole already exists nearby (within ~20 meters)
        THRESHOLD = 0.0002  # roughly 20 meters
        existing = Pathole.query.filter(
            Pathole.latitude.between(lat - THRESHOLD, lat + THRESHOLD),
            Pathole.longitude.between(lng - THRESHOLD, lng + THRESHOLD),
            Pathole.is_active == True
        ).first()

        if existing:
            # Increase report count and confidence
            existing.report_count = (existing.report_count or 1) + 1
            existing.confidence = min(1.0, (existing.confidence or 0.5) + 0.1)
            # Upgrade severity if reported many times
            if existing.report_count >= 10:
                existing.severity = 'high'
            elif existing.report_count >= 5:
                existing.severity = 'medium'
            existing.updated_at = datetime.now(timezone.utc)
            db.session.commit()
            return jsonify({
                'status': 'success',
                'message': 'Existing pathole report updated',
                'pathole': existing.to_dict()
            })

        # Determine severity from explicit input or acceleration peak
        severity_input = str(data.get('severity', '')).lower().strip()
        accel_peak = float(data.get('accel_peak', 0) or 0)

        if severity_input in ['low', 'medium', 'high']:
            severity = severity_input
        elif accel_peak >= 25:
            severity = 'high'
        elif accel_peak >= 15:
            severity = 'medium'
        else:
            severity = 'low'

        confidence_val = float(data.get('confidence', 0.6) or 0.6)

        pathole = Pathole(
            latitude=lat,
            longitude=lng,
            severity=severity,
            confidence=confidence_val,
            reported_by=str(data.get('reported_by', 'anonymous')),
            accel_peak=accel_peak
        )
        db.session.add(pathole)
        db.session.commit()

        # Cloud replicate to Turso Edge database
        sync_to_turso(
            "INSERT INTO pathole (latitude, longitude, severity, confidence, reported_by, report_count, accel_peak, created_at, updated_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [pathole.latitude, pathole.longitude, pathole.severity, pathole.confidence, pathole.reported_by, pathole.report_count, pathole.accel_peak, pathole.created_at.isoformat() if pathole.created_at else None, pathole.updated_at.isoformat() if pathole.updated_at else None, 1]
        )

        return jsonify({
            'status': 'success',
            'message': 'New pathole reported',
            'pathole': pathole.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"[PathPulse] Primary save failed ({e}), attempting auto-recovery...")
        try:
            with app.app_context():
                db.create_all()
            db.session.add(pathole)
            db.session.commit()

            sync_to_turso(
                "INSERT INTO pathole (latitude, longitude, severity, confidence, reported_by, report_count, accel_peak, created_at, updated_at, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [pathole.latitude, pathole.longitude, pathole.severity, pathole.confidence, pathole.reported_by, pathole.report_count, pathole.accel_peak, pathole.created_at.isoformat() if pathole.created_at else None, pathole.updated_at.isoformat() if pathole.updated_at else None, 1]
            )

            return jsonify({
                'status': 'success',
                'message': 'New pathole reported (auto-recovered)',
                'pathole': pathole.to_dict()
            }), 201
        except Exception as retry_err:
            db.session.rollback()
            print(f"[PathPulse] Database save error: {e} | Recovery error: {retry_err}")
            return jsonify({'status': 'error', 'message': f'Database error: {str(e)}'}), 500



@app.route('/api/patholes/<int:pathole_id>/resolve', methods=['POST'])
def resolve_pathole(pathole_id):
    """Mark a pathole as resolved/fixed"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    pathole = Pathole.query.get_or_404(pathole_id)
    pathole.is_active = False
    pathole.updated_at = datetime.now(timezone.utc)
    db.session.commit()

    sync_to_turso("UPDATE pathole SET is_active = 0, updated_at = ? WHERE id = ?", [pathole.updated_at.isoformat() if pathole.updated_at else None, pathole_id])

    return jsonify({'status': 'success', 'message': 'Pathole marked as resolved'})


@app.route('/api/patholes/<int:pathole_id>/edit', methods=['POST', 'PUT'])
def edit_pathole(pathole_id):
    """Update pathole details"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    pathole = Pathole.query.get_or_404(pathole_id)
    data = request.get_json() or {}
    
    try:
        pathole.latitude = float(data['latitude'])
        pathole.longitude = float(data['longitude'])
        pathole.severity = data.get('severity', pathole.severity)
        pathole.confidence = float(data.get('confidence', pathole.confidence))
        pathole.is_active = bool(data.get('is_active', pathole.is_active))
        pathole.updated_at = datetime.now(timezone.utc)
        
        db.session.commit()

        sync_to_turso(
            "UPDATE pathole SET latitude = ?, longitude = ?, severity = ?, confidence = ?, is_active = ?, updated_at = ? WHERE id = ?",
            [pathole.latitude, pathole.longitude, pathole.severity, pathole.confidence, int(pathole.is_active), pathole.updated_at.isoformat() if pathole.updated_at else None, pathole_id]
        )

        return jsonify({
            'status': 'success', 
            'message': 'Pathole updated successfully',
            'pathole': pathole.to_dict()
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/api/patholes/<int:pathole_id>/delete', methods=['DELETE', 'POST'])
def delete_pathole(pathole_id):
    """Permanently delete a pathole report"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    pathole = Pathole.query.get_or_404(pathole_id)
    db.session.delete(pathole)
    db.session.commit()

    sync_to_turso("DELETE FROM pathole WHERE id = ?", [pathole_id])

    return jsonify({'status': 'success', 'message': 'Pathole deleted permanently'})


# ─── Data Export Endpoints (Admin Only) ────────────────────────────────────────

def _get_filtered_patholes(status_filter):
    """Helper to query patholes by status filter ('active', 'resolved', 'all')"""
    if status_filter == 'active':
        return Pathole.query.filter_by(is_active=True).order_by(Pathole.id.asc()).all()
    elif status_filter == 'resolved':
        return Pathole.query.filter_by(is_active=False).order_by(Pathole.id.asc()).all()
    return Pathole.query.order_by(Pathole.id.asc()).all()


@app.route('/admin/export/csv', methods=['GET'])
def export_patholes_csv():
    """Export collected pothole dataset to government-ready CSV format"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Forbidden: Admin authentication required.'}), 403

    status_filter = request.args.get('status', 'all')
    patholes = _get_filtered_patholes(status_filter)

    if not patholes:
        return jsonify({'status': 'error', 'message': 'No pothole data available for export.'}), 404

    output = io.StringIO()
    writer = csv.writer(output)

    # Official government-friendly column headers
    headers = [
        'Pothole ID',
        'Latitude',
        'Longitude',
        'Severity',
        'Confidence',
        'Detection Date',
        'Detection Time',
        'Report Count',
        'Accelerometer Peak (m/s²)',
        'Verification Status',
        'Reported By'
    ]
    writer.writerow(headers)

    for p in patholes:
        det_date = p.created_at.strftime('%Y-%m-%d') if p.created_at else 'N/A'
        det_time = p.created_at.strftime('%H:%M:%S UTC') if p.created_at else 'N/A'
        status_str = 'Active' if p.is_active else 'Resolved'
        accel_str = f"{p.accel_peak:.2f}" if p.accel_peak is not None else 'N/A'
        conf_str = f"{(p.confidence * 100):.1f}%" if p.confidence is not None else 'N/A'

        writer.writerow([
            p.id,
            f"{p.latitude:.6f}",
            f"{p.longitude:.6f}",
            (p.severity or 'medium').capitalize(),
            conf_str,
            det_date,
            det_time,
            p.report_count or 1,
            accel_str,
            status_str,
            p.reported_by or 'Anonymous'
        ])

    today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    filename = f"PathPulse_Pothole_Data_{today_str}.csv"

    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )


@app.route('/admin/export/excel', methods=['GET'])
def export_patholes_excel():
    """Export collected pothole dataset to formatted Excel (.xlsx) file"""
    if not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Forbidden: Admin authentication required.'}), 403

    status_filter = request.args.get('status', 'all')
    patholes = _get_filtered_patholes(status_filter)

    if not patholes:
        return jsonify({'status': 'error', 'message': 'No pothole data available for export.'}), 404

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'status': 'error', 'message': 'openpyxl library not installed.'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Pothole Data"

    # Title Banner Row (A1:K1)
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "PathPulse AI — Smart Pothole Detection & Road Quality Report"
    title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='064E3B', end_color='064E3B', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Column Headers (Row 2)
    headers = [
        'Pothole ID',
        'Latitude',
        'Longitude',
        'Severity',
        'Confidence',
        'Detection Date',
        'Detection Time',
        'Report Count',
        'Accelerometer Peak (m/s²)',
        'Verification Status',
        'Reported By'
    ]
    ws.append(headers)
    ws.row_dimensions[2].height = 26

    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data Rows
    for p in patholes:
        det_date = p.created_at.strftime('%Y-%m-%d') if p.created_at else 'N/A'
        det_time = p.created_at.strftime('%H:%M:%S UTC') if p.created_at else 'N/A'
        status_str = 'Active' if p.is_active else 'Resolved'
        accel_val = round(p.accel_peak, 2) if p.accel_peak is not None else 'N/A'
        conf_str = f"{(p.confidence * 100):.1f}%" if p.confidence is not None else 'N/A'

        row = [
            p.id,
            float(f"{p.latitude:.6f}"),
            float(f"{p.longitude:.6f}"),
            (p.severity or 'medium').capitalize(),
            conf_str,
            det_date,
            det_time,
            p.report_count or 1,
            accel_val,
            status_str,
            p.reported_by or 'Anonymous'
        ]
        ws.append(row)

    # Format Data Rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        ws.row_dimensions[row[0].row].height = 20
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name='Segoe UI', size=10)
            if cell.column in [1, 8]:  # ID, Report Count
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.column in [2, 3]:  # Lat, Long
                cell.number_format = '0.000000'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif cell.column in [4, 5, 6, 7, 9, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                continue  # ignore merged title banner for column width
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    # Enable filters on headers and freeze panes at row 3
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    filename = f"PathPulse_Pothole_Data_{today_str}.xlsx"

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate"
        }
    )


@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get system statistics with breakdown details for visualization charts"""
    total = Pathole.query.count()
    active = Pathole.query.filter_by(is_active=True).count()
    resolved = total - active
    high_severity = Pathole.query.filter_by(is_active=True, severity='high').count()

    # Severity distribution
    low_count = Pathole.query.filter_by(is_active=True, severity='low').count()
    med_count = Pathole.query.filter_by(is_active=True, severity='medium').count()
    high_count = Pathole.query.filter_by(is_active=True, severity='high').count()

    # Timeline stats (last 7 days)
    from datetime import timedelta
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)

    daily_reports = db.session.query(
        db.func.date(Pathole.created_at).label('date_str'),
        db.func.count(Pathole.id).label('cnt')
    ).filter(Pathole.created_at >= seven_days_ago)\
     .group_by(db.func.date(Pathole.created_at))\
     .order_by('date_str').all()

    timeline = {str(row.date_str): row.cnt for row in daily_reports if row.date_str is not None}

    chart_labels = []
    chart_data = []
    for i in range(6, -1, -1):
        d = (datetime.now(timezone.utc) - timedelta(days=i))
        d_str = d.strftime('%Y-%m-%d')
        chart_labels.append(d.strftime('%a'))  # e.g. Mon, Tue
        chart_data.append(timeline.get(d_str, 0))

    return jsonify({
        'status': 'success',
        'stats': {
            'total_reported': total,
            'active_patholes': active,
            'resolved': resolved,
            'high_severity': high_severity,
            'severity_distribution': {
                'low': low_count,
                'medium': med_count,
                'high': high_count
            },
            'weekly_timeline': {
                'labels': chart_labels,
                'data': chart_data
            }
        }
    })


@app.route('/manifest.json')
def manifest():
    return send_from_directory(basedir, 'manifest.json')

@app.route('/service-worker.js')
def service_worker():
    return send_from_directory(basedir, 'service-worker.js')

# ─── Initialize Database ───────────────────────────────────────────────────────

try:
    with app.app_context():
        db.create_all()
except Exception as e:
    print(f"[PathPulse] Database initialization warning: {e}")

if __name__ == '__main__':
    app.run(debug=True, port=5000)


